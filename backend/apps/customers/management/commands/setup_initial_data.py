import os
import openpyxl
from datetime import datetime
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from apps.customers.models import Customer, Region
from apps.activities.models import ActivityType


class Command(BaseCommand):
    help = 'Set up initial data: user, activity types, and import customers from Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel-file',
            type=str,
            help='Path to the Excel file with customer data'
        )

    def handle(self, *args, **options):
        self.stdout.write('Setting up initial data...\n')

        # Create John Lischer user
        self.create_user()

        # Create activity types
        self.create_activity_types()

        # Import customers if file provided
        excel_file = options.get('excel_file')
        if excel_file:
            self.import_customers(excel_file)
        else:
            # Try default location
            default_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(
                    os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
                ))),
                'Customer_List_By_Region_CRM_Dropdowns.xlsx'
            )
            if os.path.exists(default_path):
                self.import_customers(default_path)
            else:
                self.stdout.write(self.style.WARNING('No Excel file found for import'))

        self.stdout.write(self.style.SUCCESS('\nInitial data setup complete!'))

    def create_user(self):
        """Create the John Lischer user."""
        user, created = User.objects.get_or_create(
            username='jlischer',
            defaults={
                'email': 'john.lischer@example.com',
                'first_name': 'John',
                'last_name': 'Lischer',
                'is_staff': True,
            }
        )

        if created:
            user.set_password('crm2024!')
            user.save()
            self.stdout.write(self.style.SUCCESS(f'Created user: {user.username}'))
        else:
            self.stdout.write(f'User already exists: {user.username}')

    def create_activity_types(self):
        """Create default activity types."""
        activity_types = [
            {
                'name': 'phone_call',
                'display_name': 'Phone Call',
                'icon': 'phone',
                'color': '#3B82F6',
                'is_calendar_event': False,
                'default_duration_minutes': 15,
                'sort_order': 1,
            },
            {
                'name': 'text_message',
                'display_name': 'Text Message',
                'icon': 'message-square',
                'color': '#10B981',
                'is_calendar_event': False,
                'default_duration_minutes': 5,
                'sort_order': 2,
            },
            {
                'name': 'email',
                'display_name': 'Email',
                'icon': 'mail',
                'color': '#6366F1',
                'is_calendar_event': False,
                'default_duration_minutes': 10,
                'sort_order': 3,
            },
            {
                'name': 'meeting',
                'display_name': 'Meeting',
                'icon': 'calendar',
                'color': '#F59E0B',
                'is_calendar_event': True,
                'default_duration_minutes': 60,
                'sort_order': 4,
            },
            {
                'name': 'cold_call',
                'display_name': 'Cold Call',
                'icon': 'phone-outgoing',
                'color': '#EF4444',
                'is_calendar_event': False,
                'default_duration_minutes': 10,
                'sort_order': 5,
            },
            {
                'name': 'card_drop',
                'display_name': 'Card Drop',
                'icon': 'id-card',
                'color': '#8B5CF6',
                'is_calendar_event': False,
                'default_duration_minutes': 5,
                'sort_order': 6,
            },
        ]

        for at_data in activity_types:
            at, created = ActivityType.objects.get_or_create(
                name=at_data['name'],
                defaults=at_data
            )
            if created:
                self.stdout.write(f'Created activity type: {at.display_name}')

    def import_customers(self, file_path):
        """Import customers from Excel file."""
        self.stdout.write(f'\nImporting customers from: {file_path}')

        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error reading Excel file: {e}'))
            return

        user = User.objects.filter(username='jlischer').first()
        total_created = 0
        total_skipped = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self.stdout.write(f'\nProcessing sheet: {sheet_name}')

            # Get or create region
            region, _ = Region.objects.get_or_create(
                name=sheet_name,
                defaults={'description': f'Region from import: {sheet_name}'}
            )

            # Get headers from first row
            headers = [cell.value for cell in ws[1]]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                row_data = dict(zip(headers, row))

                # Extract customer name (first column)
                business_name = row_data.get('Customer') or row_data.get(headers[0])
                if not business_name:
                    continue

                # Skip if already exists
                if Customer.objects.filter(business_name=business_name).exists():
                    total_skipped += 1
                    continue

                # Parse location/address
                location = row_data.get('Location') or row_data.get(headers[1]) or ''

                # Extract contact and phone
                contact = None
                phone = None
                secondary_phone = None
                email = None

                # The Excel structure varies by sheet, so we need to handle it flexibly
                for key, value in row_data.items():
                    if not value:
                        continue
                    value_str = str(value)
                    key_lower = str(key).lower() if key else ''

                    if 'contact' in key_lower or key_lower == 'Contact':
                        contact = value_str
                    elif 'email' in key_lower or '@' in value_str:
                        email = value_str
                    elif 'number' in key_lower or 'phone' in key_lower:
                        if not phone:
                            phone = value_str
                        elif not secondary_phone:
                            secondary_phone = value_str

                # Try to parse phone numbers from various columns
                for i, val in enumerate(row):
                    if val and isinstance(val, str):
                        # Check if it looks like a phone number
                        digits = ''.join(c for c in val if c.isdigit())
                        if len(digits) >= 10 and not phone:
                            phone = val
                        elif len(digits) >= 10 and not secondary_phone:
                            secondary_phone = val

                # Get fleet description/notes from later columns
                fleet_desc = ''
                for i, val in enumerate(row):
                    if val and i > 4:  # After the main contact columns
                        if isinstance(val, str) and len(val) > 10 and not val.startswith('http'):
                            if 'fleet' in str(headers[i]).lower() if i < len(headers) else False:
                                fleet_desc = val
                            elif not fleet_desc and len(val) > 20:
                                fleet_desc = val

                # Handle call dates
                last_call = None
                next_call = None
                for key, value in row_data.items():
                    if value and isinstance(value, datetime):
                        key_lower = str(key).lower() if key else ''
                        if 'call date' in key_lower or 'last' in key_lower:
                            last_call = value.date()
                        elif 'next' in key_lower:
                            next_call = value.date()

                # Create customer
                Customer.objects.create(
                    business_name=business_name,
                    bill_to_address=location,
                    primary_contact=contact or '',
                    main_email=email or '',
                    main_phone=phone or '',
                    secondary_phone=secondary_phone or '',
                    fleet_description=fleet_desc,
                    region=region,
                    last_call_date=last_call,
                    next_call_date=next_call,
                    created_by=user,
                )
                total_created += 1

        self.stdout.write(self.style.SUCCESS(
            f'\nImport complete: {total_created} created, {total_skipped} skipped (duplicates)'
        ))
