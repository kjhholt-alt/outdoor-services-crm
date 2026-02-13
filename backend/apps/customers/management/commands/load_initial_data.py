import os
import openpyxl
from datetime import datetime
from django.core.management.base import BaseCommand
from apps.customers.models import Customer, Region
from apps.activities.models import ActivityType


class Command(BaseCommand):
    help = 'Load initial customer data from Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--force',
            action='store_true',
            help='Delete existing customers and reimport all data',
        )

    def handle(self, *args, **options):
        force = options.get('force', False)

        # Check if data already exists
        existing_count = Customer.objects.count()
        if existing_count > 0:
            if force:
                self.stdout.write(self.style.WARNING(
                    f'Force flag set. Deleting {existing_count} existing customers...'
                ))
                Customer.objects.all().delete()
                Region.objects.all().delete()
                self.stdout.write(self.style.SUCCESS('Existing data deleted.'))
            else:
                self.stdout.write(self.style.WARNING(
                    f'Found {existing_count} existing customers. Skipping import.\n'
                    f'Use --force to delete existing data and reimport.'
                ))
                return

        self.stdout.write('Loading initial data...\n')

        # Create activity types first
        self.create_activity_types()

        # Import customers from Excel
        excel_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(
                os.path.dirname(os.path.dirname(__file__))
            ))),
            'initial_data.xlsx'
        )

        if os.path.exists(excel_path):
            self.import_customers(excel_path)
        else:
            self.stdout.write(self.style.ERROR(f'Excel file not found at: {excel_path}'))

        self.stdout.write(self.style.SUCCESS('\nInitial data load complete!'))

    def create_activity_types(self):
        """Create default activity types."""
        activity_types = [
            {
                'name': 'phone_call',
                'display_name': 'Phone Call',
                'icon': 'phone',
                'color': '#3B82F6',
                'is_calendar_event': False,
                'default_duration_minutes': 15,
                'sort_order': 1,
            },
            {
                'name': 'text_message',
                'display_name': 'Text Message',
                'icon': 'message-square',
                'color': '#10B981',
                'is_calendar_event': False,
                'default_duration_minutes': 5,
                'sort_order': 2,
            },
            {
                'name': 'email',
                'display_name': 'Email',
                'icon': 'mail',
                'color': '#6366F1',
                'is_calendar_event': False,
                'default_duration_minutes': 10,
                'sort_order': 3,
            },
            {
                'name': 'meeting',
                'display_name': 'Meeting',
                'icon': 'calendar',
                'color': '#F59E0B',
                'is_calendar_event': True,
                'default_duration_minutes': 60,
                'sort_order': 4,
            },
            {
                'name': 'cold_call',
                'display_name': 'Cold Call',
                'icon': 'phone-outgoing',
                'color': '#EF4444',
                'is_calendar_event': False,
                'default_duration_minutes': 10,
                'sort_order': 5,
            },
            {
                'name': 'card_drop',
                'display_name': 'Card Drop',
                'icon': 'id-card',
                'color': '#8B5CF6',
                'is_calendar_event': False,
                'default_duration_minutes': 5,
                'sort_order': 6,
            },
        ]

        for at_data in activity_types:
            at, created = ActivityType.objects.get_or_create(
                name=at_data['name'],
                defaults=at_data
            )
            if created:
                self.stdout.write(f'Created activity type: {at.display_name}')

    def import_customers(self, file_path):
        """Import customers from Excel file."""
        self.stdout.write(f'\nImporting customers from: {file_path}')

        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error reading Excel file: {e}'))
            return

        total_created = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self.stdout.write(f'\nProcessing sheet: {sheet_name}')

            # Get or create region
            region, _ = Region.objects.get_or_create(
                name=sheet_name,
                defaults={'description': f'Region: {sheet_name}'}
            )

            # Get headers from first row
            headers = [cell.value for cell in ws[1]]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                row_data = dict(zip(headers, row))

                # Extract customer name (first column)
                business_name = row_data.get('Customer') or row_data.get(headers[0])
                if not business_name:
                    continue

                # Parse location/address
                location = row_data.get('Location') or row_data.get(headers[1]) or ''

                # Extract contact and phone
                contact = None
                phone = None
                secondary_phone = None
                email = None

                for key, value in row_data.items():
                    if not value:
                        continue
                    value_str = str(value)
                    key_lower = str(key).lower() if key else ''

                    if 'contact' in key_lower:
                        contact = value_str
                    elif 'email' in key_lower or '@' in value_str:
                        email = value_str
                    elif 'number' in key_lower or 'phone' in key_lower:
                        if not phone:
                            phone = value_str
                        elif not secondary_phone:
                            secondary_phone = value_str

                # Try to parse phone numbers from various columns
                for i, val in enumerate(row):
                    if val and isinstance(val, str):
                        digits = ''.join(c for c in val if c.isdigit())
                        if len(digits) >= 10 and not phone:
                            phone = val
                        elif len(digits) >= 10 and not secondary_phone:
                            secondary_phone = val

                # Get fleet description from later columns
                fleet_desc = ''
                for i, val in enumerate(row):
                    if val and i > 4:
                        if isinstance(val, str) and len(val) > 10 and not val.startswith('http'):
                            if i < len(headers) and headers[i] and 'fleet' in str(headers[i]).lower():
                                fleet_desc = val
                            elif not fleet_desc and len(val) > 20:
                                fleet_desc = val

                # Handle call dates
                last_call = None
                next_call = None
                for key, value in row_data.items():
                    if value and isinstance(value, datetime):
                        key_lower = str(key).lower() if key else ''
                        if 'call date' in key_lower or 'last' in key_lower:
                            last_call = value.date()
                        elif 'next' in key_lower:
                            next_call = value.date()

                # Create customer
                try:
                    Customer.objects.create(
                        business_name=business_name,
                        bill_to_address=location,
                        primary_contact=contact or '',
                        main_email=email or '',
                        main_phone=phone or '',
                        secondary_phone=secondary_phone or '',
                        fleet_description=fleet_desc,
                        region=region,
                        last_call_date=last_call,
                        next_call_date=next_call,
                    )
                    total_created += 1
                except Exception as e:
                    self.stdout.write(self.style.ERROR(
                        f'  Error creating "{business_name}": {e}'
                    ))

        self.stdout.write(self.style.SUCCESS(
            f'\n========================================\n'
            f'Successfully imported {total_created} customers\n'
            f'========================================'
        ))
