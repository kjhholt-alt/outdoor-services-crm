import openpyxl
import csv
from io import BytesIO, StringIO
from django.db import transaction
from apps.customers.models import Customer, Region


class ImportService:
    """Service for importing customer data from Excel and CSV files."""

    REQUIRED_FIELDS = ['business_name']
    FIELD_MAPPINGS = {
        'customer': 'business_name',
        'company': 'business_name',
        'company_name': 'business_name',
        'name': 'business_name',
        'location': 'bill_to_address',
        'address': 'bill_to_address',
        'contact': 'primary_contact',
        'contact_name': 'primary_contact',
        'phone': 'main_phone',
        'phone_number': 'main_phone',
        'number': 'main_phone',
        'email': 'main_email',
        'fax': 'fax',
        'fleet': 'fleet_description',
        'fleet_description': 'fleet_description',
        'notes': 'fleet_description',
        'region': 'region',
        'city': 'city',
        'state': 'state',
        'zip': 'zip_code',
        'zip_code': 'zip_code',
    }

    def __init__(self, file_obj, file_type='xlsx'):
        self.file_obj = file_obj
        self.file_type = file_type
        self.errors = []
        self.warnings = []

    def read_excel(self):
        """Read data from Excel file."""
        wb = openpyxl.load_workbook(self.file_obj)
        all_data = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1] if cell.value]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                row_data = dict(zip(headers, row))
                row_data['_sheet'] = sheet_name  # Track source sheet
                all_data.append(row_data)

        return all_data

    def read_csv(self):
        """Read data from CSV file."""
        content = self.file_obj.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8')

        reader = csv.DictReader(StringIO(content))
        return list(reader)

    def preview(self, field_mapping=None):
        """Preview import data with optional field mapping."""
        if self.file_type == 'xlsx':
            data = self.read_excel()
        else:
            data = self.read_csv()

        if not data:
            return {'error': 'No data found in file', 'rows': []}

        # Get all unique headers
        headers = set()
        for row in data:
            headers.update(row.keys())
        headers.discard('_sheet')

        # Auto-map fields if no mapping provided
        if not field_mapping:
            field_mapping = {}
            for header in headers:
                normalized = header.lower().strip().replace(' ', '_')
                if normalized in self.FIELD_MAPPINGS:
                    field_mapping[header] = self.FIELD_MAPPINGS[normalized]

        # Preview first 10 rows
        preview_rows = []
        for row in data[:10]:
            mapped_row = {}
            for src, dest in field_mapping.items():
                if src in row:
                    mapped_row[dest] = row[src]
            if '_sheet' in row:
                mapped_row['_region'] = row['_sheet']
            preview_rows.append(mapped_row)

        return {
            'headers': list(headers),
            'field_mapping': field_mapping,
            'row_count': len(data),
            'preview_rows': preview_rows,
            'available_fields': [
                'business_name', 'bill_to_address', 'city', 'state', 'zip_code',
                'primary_contact', 'main_email', 'main_phone', 'secondary_phone',
                'fax', 'fleet_description', 'region'
            ]
        }

    def execute(self, field_mapping, user, duplicate_action='skip'):
        """
        Execute the import.

        duplicate_action: 'skip' | 'update' | 'create_new'
        """
        if self.file_type == 'xlsx':
            data = self.read_excel()
        else:
            data = self.read_csv()

        results = {
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': []
        }

        with transaction.atomic():
            for i, row in enumerate(data, start=2):  # Start at 2 (header is row 1)
                try:
                    # Map fields
                    customer_data = {}
                    for src, dest in field_mapping.items():
                        if src in row and row[src]:
                            value = row[src]
                            # Clean string values
                            if isinstance(value, str):
                                value = value.strip()
                            customer_data[dest] = value

                    # Use sheet name as region if not mapped
                    if 'region' not in customer_data and '_sheet' in row:
                        customer_data['region'] = row['_sheet']

                    # Validate required fields
                    if 'business_name' not in customer_data or not customer_data['business_name']:
                        results['errors'].append(f"Row {i}: Missing business name")
                        continue

                    # Handle region
                    region = None
                    region_name = customer_data.pop('region', None)
                    if region_name:
                        region, _ = Region.objects.get_or_create(
                            name=region_name,
                            defaults={'description': f'Imported region: {region_name}'}
                        )

                    # Check for duplicates
                    existing = Customer.objects.filter(
                        business_name__iexact=customer_data['business_name']
                    ).first()

                    if existing:
                        if duplicate_action == 'skip':
                            results['skipped'] += 1
                            continue
                        elif duplicate_action == 'update':
                            for key, value in customer_data.items():
                                setattr(existing, key, value)
                            if region:
                                existing.region = region
                            existing.save()
                            results['updated'] += 1
                            continue

                    # Create new customer
                    customer_data['region'] = region
                    customer_data['created_by'] = user
                    Customer.objects.create(**customer_data)
                    results['created'] += 1

                except Exception as e:
                    results['errors'].append(f"Row {i}: {str(e)}")

        return results


def export_customers_to_excel(queryset):
    """Export customers to Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Headers
    headers = [
        'Business Name', 'Address', 'City', 'State', 'Zip Code',
        'Contact', 'Email', 'Phone', 'Fax', 'Fleet Description',
        'Region', 'Last Call Date', 'Next Call Date'
    ]
    ws.append(headers)

    # Data
    for customer in queryset:
        ws.append([
            customer.business_name,
            customer.bill_to_address,
            customer.city,
            customer.state,
            customer.zip_code,
            customer.primary_contact,
            customer.main_email,
            customer.main_phone,
            customer.fax,
            customer.fleet_description,
            customer.region.name if customer.region else '',
            customer.last_call_date.isoformat() if customer.last_call_date else '',
            customer.next_call_date.isoformat() if customer.next_call_date else '',
        ])

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_customers_to_csv(queryset):
    """Export customers to CSV file."""
    output = StringIO()
    writer = csv.writer(output)

    # Headers
    headers = [
        'Business Name', 'Address', 'City', 'State', 'Zip Code',
        'Contact', 'Email', 'Phone', 'Fax', 'Fleet Description',
        'Region', 'Last Call Date', 'Next Call Date'
    ]
    writer.writerow(headers)

    # Data
    for customer in queryset:
        writer.writerow([
            customer.business_name,
            customer.bill_to_address,
            customer.city,
            customer.state,
            customer.zip_code,
            customer.primary_contact,
            customer.main_email,
            customer.main_phone,
            customer.fax,
            customer.fleet_description,
            customer.region.name if customer.region else '',
            customer.last_call_date.isoformat() if customer.last_call_date else '',
            customer.next_call_date.isoformat() if customer.next_call_date else '',
        ])

    output.seek(0)
    return output
